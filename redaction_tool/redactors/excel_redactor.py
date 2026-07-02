"""
Excel (.xlsx) Redactor — uses openpyxl.

Iterates all sheets, rows, and cells. Replaces matched text in cell values.
Formula cells: the formula string is also scanned (sensitive data sometimes
appears in formula arguments). Cell formatting is preserved.
"""

import re
from pathlib import Path


def redact_excel(
    input_path: str,
    output_path: str,
    patterns: list,
    custom_terms: list,
    replacement: str = "[REDACTED]",
    case_sensitive: bool = False,
) -> tuple:
    """Returns (total_count, audit_entries)."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel redaction.\n"
            "Install with:  pip install openpyxl"
        )

    flags = 0 if case_sensitive else re.IGNORECASE

    compiled = []
    for regex, label in patterns:
        try:
            compiled.append((re.compile(regex, flags), label))
        except re.error:
            pass
    for kw in custom_terms:
        kw = kw.strip()
        if kw:
            try:
                compiled.append((re.compile(re.escape(kw), flags), "Custom Keyword"))
            except re.error:
                pass

    # data_only=False preserves formulas; keep_vba=False drops macros
    wb = openpyxl.load_workbook(input_path, data_only=False, keep_vba=False)
    total = 0
    audit = []
    file_name = Path(input_path).name

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                new_val = val
                for pat, label in compiled:
                    matches = list(pat.finditer(new_val))
                    if not matches:
                        continue
                    for m in matches:
                        audit.append({
                            "File": file_name,
                            "Page": f"Sheet: {sheet_name}",
                            "Category": label,
                            "Original Term": (m.group()[:45] + "…") if len(m.group()) > 45 else m.group(),
                            "Count": 1,
                        })
                    new_val, n = pat.subn(replacement, new_val)
                    total += n
                if new_val != val:
                    # Preserve numeric cells that became strings after redaction
                    cell.value = new_val

    wb.save(output_path)
    return total, audit
