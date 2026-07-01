"""
Cross-check SOA (actual) against RPS (scheduled) for the same loan.

Matches a Statement of Account to its Repayment Schedule by Agreement No and
reports where actual servicing deviated from the contractual plan:
  - Term_Checks            : static terms (loan amount, rate, tenure, EMI, disb date)
  - Instalment_Comparison  : per instalment, scheduled vs actual (amount / date / DPD)
  - Reconciliation_Summary : one row per agreement with deviation counts + result
  - Unmatched              : agreements present on only one side

Usage (library): build SOA dicts via extract_soa.extract_loan and RPS dicts via
extract_rps.extract_rps, then call reconcile_jobs([...]).
"""
import re
import datetime as dt

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from extract_soa import style_header, autosize, to_num, TITLE_FONT, STATUS_FILL

TOL = 1.0


def classify(pdf_path):
    """Identify a PDF as 'soa', 'rps', or 'unknown' from its first page."""
    with pdfplumber.open(pdf_path) as pdf:
        t = pdf.pages[0].extract_text() or ""
    if re.search(r"Repayment Schedule of", t):
        return "rps"
    if re.search(r"Loan account statement|Statement of Account", t, re.I):
        return "soa"
    return "unknown"


def _norm_date(s):
    if isinstance(s, dt.date):
        return s
    if not s:
        return None
    for fmt in ("%d-%b-%Y", "%d-%b-%y", "%d-%B-%Y"):
        try:
            return dt.datetime.strptime(s.strip(), fmt).date()
        except (ValueError, AttributeError):
            continue
    return None


def reconcile_pair(soa, rps):
    agr = soa["master"].get("Agreement No") or rps["master"].get("Agreement No")
    term_checks = []

    def tc(item, rv, sv, ok):
        term_checks.append([agr, item, rv, sv, "MATCH" if ok else "MISMATCH"])

    num_fields = [("Loan Amount (Rs)", TOL), ("Annualised Interest Rate %", 0.01),
                  ("Loan Tenure (Months)", 0), ("First Instalment Amount (Rs)", TOL)]
    for f, tol in num_fields:
        rv = rps["master"].get(f)
        if f == "First Instalment Amount (Rs)" and rv is None:
            rv = rps["master"].get("First Instalment / EMI Amount (Rs)")
        sv = soa["master"].get(f)
        rn, sn = to_num(rv), to_num(sv)
        tc(f, rv, sv, rn is not None and sn is not None and abs(rn - sn) <= tol)
    rv, sv = rps["master"].get("Disbursement Date"), soa["master"].get("Disbursement Date")
    tc("Disbursement Date", rv, sv, _norm_date(rv) == _norm_date(sv) and _norm_date(rv) is not None)

    soa_by_n = {d["Instalment"]: d for d in soa["dpd_rows"]}
    instalments = []
    counts = {"matched": 0, "amount": 0, "late": 0, "unpaid": 0, "missing": 0}
    for r in rps["schedule"]:
        n = r["Instalment Number"]
        principal, interest = r.get("Principal"), r.get("Interest")
        # scheduled EMI = Principal + Interest; "Instalment Balance" is an
        # outstanding/running balance figure, not the instalment amount.
        sched_emi = (principal + interest) if principal is not None and interest is not None else None
        sd = soa_by_n.get(n)
        soa_amt = soa_due = soa_paid = dpd = None
        if not sd:
            status = "NOT IN SOA"
            counts["missing"] += 1
        else:
            soa_amt, soa_due = sd["Amount"], sd["Due Date"]
            soa_paid, dpd = sd["Paid Date"], sd["DPD"]
            if soa_amt is not None and sched_emi is not None and abs(soa_amt - sched_emi) > TOL:
                status = "AMOUNT MISMATCH"; counts["amount"] += 1
            elif soa_paid is None:
                status = "UNPAID"; counts["unpaid"] += 1
            elif dpd and dpd > 0:
                status = f"PAID LATE ({dpd}d)"; counts["late"] += 1
            else:
                status = "MATCH"; counts["matched"] += 1
        instalments.append([agr, n, r["Instalment Date"], sched_emi, r["Principal"],
                            r["Interest"], r["Closing Balance"],
                            str(soa_due) if soa_due else "", soa_amt,
                            str(soa_paid) if soa_paid else "", dpd if dpd else 0, status])

    term_mismatch = sum(1 for c in term_checks if c[4] == "MISMATCH")
    deviations = term_mismatch + counts["amount"] + counts["late"] + counts["unpaid"]
    summary = {
        "agreement": agr, "term_mismatch": term_mismatch,
        "scheduled": len(rps["schedule"]), **counts,
        "result": "EXCEPTION" if deviations else "RECONCILED",
    }
    return {"agreement": agr, "term_checks": term_checks,
            "instalments": instalments, "summary": summary}


def reconcile_jobs(soas, rpss):
    soa_by = {s["master"].get("Agreement No"): s for s in soas}
    rps_by = {r["master"].get("Agreement No"): r for r in rpss}
    common = [a for a in soa_by if a in rps_by]
    pairs = [reconcile_pair(soa_by[a], rps_by[a]) for a in common]
    soa_only = [a for a in soa_by if a not in rps_by]
    rps_only = [a for a in rps_by if a not in soa_by]
    return pairs, soa_only, rps_only


def write_reconciliation_workbook(path, pairs, soa_only, rps_only):
    wb = Workbook()

    ws = wb.active; ws.title = "Reconciliation_Summary"
    ws["A1"] = "SOA (ACTUAL) vs RPS (SCHEDULED) RECONCILIATION"; ws["A1"].font = TITLE_FONT
    ws.append([])
    cols = ["Agreement No", "Term Mismatches", "Scheduled Instalments", "Matched",
            "Amount Mismatch", "Paid Late", "Unpaid", "Not in SOA", "Result"]
    ws.append(cols); style_header(ws, 3, len(cols))
    for p in pairs:
        s = p["summary"]
        ws.append([s["agreement"], s["term_mismatch"], s["scheduled"], s["matched"],
                   s["amount"], s["late"], s["unpaid"], s["missing"], s["result"]])
        fill = "C6EFCE" if s["result"] == "RECONCILED" else "FFC7CE"
        ws.cell(ws.max_row, len(cols)).fill = PatternFill("solid", fgColor=fill)
    autosize(ws)

    ws = wb.create_sheet("Term_Checks")
    cols = ["Agreement No", "Term", "RPS (Scheduled)", "SOA (Actual)", "Status"]
    ws.append(cols); style_header(ws, 1, len(cols))
    for p in pairs:
        for c in p["term_checks"]:
            ws.append(c)
            ws.cell(ws.max_row, 5).fill = PatternFill(
                "solid", fgColor=STATUS_FILL["PASS"] if c[4] == "MATCH" else STATUS_FILL["FAIL"])
    ws.freeze_panes = "A2"; autosize(ws)

    ws = wb.create_sheet("Instalment_Comparison")
    cols = ["Agreement No", "Inst No", "Sched Date", "Sched EMI", "Sched Principal",
            "Sched Interest", "Sched Closing", "SOA Due Date", "SOA Due Amt",
            "SOA Paid Date", "DPD", "Status"]
    ws.append(cols); style_header(ws, 1, len(cols))
    for p in pairs:
        for row in p["instalments"]:
            ws.append(row)
            if row[-1] != "MATCH":
                ws.cell(ws.max_row, len(cols)).fill = PatternFill("solid", fgColor="FFEB9C")
    ws.freeze_panes = "A2"; autosize(ws)

    if soa_only or rps_only:
        ws = wb.create_sheet("Unmatched")
        ws.append(["Agreement No", "Present in", "Note"]); style_header(ws, 1, 3)
        for a in soa_only:
            ws.append([a, "SOA only", "No matching Repayment Schedule uploaded"])
        for a in rps_only:
            ws.append([a, "RPS only", "No matching Statement of Account uploaded"])
        autosize(ws)

    wb.save(path)
