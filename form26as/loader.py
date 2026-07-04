"""Load a Form 26AS file into a uniform grid of string cells.

Supported inputs: the native caret (^) delimited TRACES ``.txt`` export, a
genuine ``.xlsx``/``.xls`` workbook, an HTML export (including TRACES "Excel"
files that are really HTML with an ``.xls``/``.xlsx`` extension), and the
password-protected ``.pdf`` download. Whatever the source, this loader returns
a single flat grid (list of rows, each a list of stripped strings). The parser
scans that grid for the header rows it needs, so it does not matter that the
different parts of the statement have different column counts.
"""

from __future__ import annotations

from pathlib import Path
from typing import List, Optional

Grid = List[List[str]]

# Signature of the OLE Compound File Binary Format, used by both a
# password-protected OOXML (.xlsx) file and a legacy pre-2007 binary (.xls)
# workbook. openpyxl cannot read either directly, and mistaking one for a
# corrupt/mislabeled workbook (and silently falling back to HTML parsing)
# used to produce an empty grid with no explanation - see EncryptedExcelError
# / LegacyExcelError below for the fix.
_OLE_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class EncryptedExcelError(Exception):
    """A password-protected Excel file was opened without (or with the wrong) password."""


class LegacyExcelError(Exception):
    """A pre-2007 binary .xls file was given; openpyxl can only read .xlsx/.xlsm."""


def _looks_like_html(head: bytes) -> bool:
    lowered = head.lower()
    return b"<html" in lowered or b"<table" in lowered or b"<!doctype html" in lowered


def _xlsx_cell_to_str(value: object) -> str:
    """Stringify a cell, keeping real date/datetime cells parseable downstream.

    Excel stores dates as datetime objects (a too-narrow column just shows
    "####" visually, the underlying value is a real date). str(datetime) adds
    a " 00:00:00" suffix that the parser's date formats don't match, silently
    downgrading the column to unparsed text - format date-only values as
    plain ISO dates instead so they parse the same as every other input type.
    """
    import datetime as _dt

    if value is None:
        return ""
    if isinstance(value, _dt.datetime):
        if value.time() == _dt.time.min:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, _dt.date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _load_xlsx(path: Path, password: Optional[str] = None) -> Grid:
    from openpyxl import load_workbook

    with open(path, "rb") as fh:
        head = fh.read(8)

    source = str(path)
    if head == _OLE_SIGNATURE:
        source = _decrypt_ole_excel(path, password)

    wb = load_workbook(filename=source, read_only=True, data_only=True)
    grid: Grid = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            grid.append([_xlsx_cell_to_str(c) for c in row])
    wb.close()
    return grid


def _decrypt_ole_excel(path: Path, password: Optional[str]):
    """Return a file-like object openpyxl can read from an OLE-format .xlsx.

    Raises EncryptedExcelError if it's encrypted and no/wrong password was
    given, or LegacyExcelError if it's a genuine pre-2007 .xls that no
    password can help with.
    """
    import io

    import msoffcrypto
    from msoffcrypto.exceptions import DecryptionError

    with open(path, "rb") as fh:
        office_file = msoffcrypto.OfficeFile(fh)
        try:
            is_encrypted = office_file.is_encrypted()
        except Exception:
            is_encrypted = False

        if not is_encrypted:
            raise LegacyExcelError(
                f"{path.name} looks like an old-format (.xls) Excel file, which this "
                "tool can't read directly. Open it in Excel and use File > Save As > "
                "'Excel Workbook (*.xlsx)', then run this tool on the new file."
            )

        if not password:
            raise EncryptedExcelError(
                f"{path.name} is a password-protected Excel file. Provide its "
                "password with --password (or enter it when prompted)."
            )

        fh.seek(0)
        office_file = msoffcrypto.OfficeFile(fh)
        try:
            office_file.load_key(password=password)
            decrypted = io.BytesIO()
            office_file.decrypt(decrypted)
        except DecryptionError as exc:
            raise EncryptedExcelError(
                f"{path.name} could not be decrypted with the given password: {exc}"
            ) from exc
        decrypted.seek(0)
        return decrypted


def _load_text(path: Path) -> Grid:
    """Split a caret (^) delimited TRACES text export into a grid.

    Falls back to tab or comma if no carets are present.
    """
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        lines = fh.read().splitlines()
    sample = "\n".join(lines[:200])
    if "^" in sample:
        delim = "^"
    elif "\t" in sample:
        delim = "\t"
    else:
        delim = ","
    return [[cell.strip() for cell in line.split(delim)] for line in lines if line.strip()]


def _load_html(path: Path) -> Grid:
    from bs4 import BeautifulSoup

    with open(path, "rb") as fh:
        raw = fh.read()
    soup = BeautifulSoup(raw, "lxml")
    grid: Grid = []
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            cells: List[str] = []
            for cell in tr.find_all(["td", "th"]):
                text = cell.get_text(separator=" ", strip=True)
                # Expand colspan so columns stay aligned across rows: the first
                # slot holds the text, the rest are blank placeholders.
                try:
                    span = int(cell.get("colspan", 1))
                except (TypeError, ValueError):
                    span = 1
                cells.append(text)
                cells.extend([""] * max(0, span - 1))
            if cells:
                grid.append(cells)
    return grid


def _split_on_wide_gaps(line: str) -> List[str]:
    """Split a text line into cells on runs of 2+ spaces (fallback for PDFs)."""
    import re

    parts = re.split(r"\s{2,}", line.strip())
    return [p.strip() for p in parts if p.strip() != ""]


def _load_pdf(path: Path, password: Optional[str] = None) -> Grid:
    """Extract TDS/TCS-style tables from a (possibly password-protected) PDF.

    Prefers pdfplumber's ruled-table extraction; if a page has no detectable
    table it falls back to splitting text lines on wide whitespace gaps.
    """
    import pdfplumber

    grid: Grid = []
    with pdfplumber.open(str(path), password=password or "") as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        grid.append(["" if c is None else str(c).replace("\n", " ").strip()
                                     for c in row])
            else:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    cells = _split_on_wide_gaps(line)
                    if cells:
                        grid.append(cells)
    return grid


def load_grid(path: str | Path, password: Optional[str] = None) -> Grid:
    """Return the file's contents as a list of rows of stripped string cells."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"No such file: {path}")

    with open(path, "rb") as fh:
        head = fh.read(4096)

    suffix = path.suffix.lower()
    if suffix in {".txt", ".csv"}:
        return _load_text(path)

    # PDF, by extension or magic bytes.
    if suffix == ".pdf" or head[:5] == b"%PDF-":
        return _load_pdf(path, password)

    # HTML masquerading as .xls/.xlsx is common from TRACES, so sniff content
    # first rather than trusting the extension.
    if _looks_like_html(head):
        return _load_html(path)

    # A caret-delimited text export sniffed by content (extension may vary).
    if b"^" in head and b"PK\x03\x04" not in head[:4]:
        return _load_text(path)

    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return _load_xlsx_or_fallback(path, password)
    if suffix in {".html", ".htm"}:
        return _load_html(path)

    # Unknown extension: try workbook, then HTML.
    return _load_xlsx_or_fallback(path, password)


def _load_xlsx_or_fallback(path: Path, password: Optional[str]) -> Grid:
    import zipfile

    from openpyxl.utils.exceptions import InvalidFileException

    try:
        return _load_xlsx(path, password)
    except (EncryptedExcelError, LegacyExcelError):
        # These mean we correctly identified the problem - surface it as-is
        # rather than obscuring it behind a doomed HTML-parse attempt.
        raise
    except (zipfile.BadZipFile, InvalidFileException):
        # Not a real workbook and not an OLE file either - most likely HTML
        # mislabeled with an .xls/.xlsx extension (common from TRACES).
        return _load_html(path)
