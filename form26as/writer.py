"""Write parsed transactions to a formatted, searchable .xlsx and/or .csv."""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path
from typing import List

from . import summary as _summary
from .parser import COLUMNS, Transaction

_NUMERIC_COLS = {
    "Total Amount Paid/Credited",
    "Total Tax Deducted/Collected",
    "Total TDS/TCS Deposited",
    "Amount Paid/Credited",
    "Tax Deducted/Collected",
    "TDS/TCS Deposited",
}
_DATE_COLS = {"Transaction Date", "Date of Booking"}


def write_csv(transactions: List[Transaction], path: str | Path) -> None:
    path = Path(path)
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(COLUMNS)
        for txn in transactions:
            row = []
            for value in txn.as_row():
                if isinstance(value, date):
                    row.append(value.isoformat())
                elif value is None:
                    row.append("")
                else:
                    row.append(value)
            writer.writerow(row)


def write_xlsx(transactions: List[Transaction], path: str | Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Form 26AS - TDS & TCS"

    header_fill = PatternFill("solid", fgColor="1F6390")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(COLUMNS)
    for col_idx, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for txn in transactions:
        ws.append(txn.as_row())

    # Apply number and date formats per column.
    for col_idx, name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        if name in _NUMERIC_COLS:
            number_format = "#,##0.00"
        elif name in _DATE_COLS:
            number_format = "DD-MMM-YYYY"
        else:
            number_format = None
        if number_format:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = number_format

    # Column widths sized to content (bounded).
    for col_idx, name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        longest = len(name)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                longest = max(longest, len(str(value)))
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), 40)

    # Freeze the header and enable AutoFilter so the sheet is searchable/sortable.
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{max(ws.max_row, 1)}"

    _write_summary_sheets(wb, transactions)

    wb.save(str(path))


def _write_summary_sheets(wb, transactions: List[Transaction]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F6390")
    header_font = Font(bold=True, color="FFFFFF")
    total_font = Font(bold=True)
    money_cols = {"Amount Paid/Credited", "Tax Deducted/Collected", "TDS/TCS Deposited"}

    grand = _summary.totals(transactions)

    def add_sheet(title, key_headers, rows):
        ws = wb.create_sheet(title=title)
        headers = list(key_headers) + [
            "Transactions",
            "Amount Paid/Credited",
            "Tax Deducted/Collected",
            "TDS/TCS Deposited",
        ]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for r in rows:
            ws.append(list(r.key) + [r.count, r.amount_paid, r.tax_deducted, r.tds_deposited])

        # Grand total row.
        pad = [""] * (len(key_headers) - 1)
        ws.append(["Total"] + pad + [grand.count, grand.amount_paid,
                                     grand.tax_deducted, grand.tds_deposited])
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col_idx).font = total_font

        # Money formatting + column widths.
        for col_idx, name in enumerate(headers, start=1):
            letter = get_column_letter(col_idx)
            if name in money_cols:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"
            longest = len(name)
            for row_idx in range(2, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    longest = max(longest, len(str(value)))
            ws.column_dimensions[letter].width = min(max(longest + 2, 12), 40)

        ws.freeze_panes = "A2"
        last = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last}{max(ws.max_row, 1)}"

    categories = {t.category for t in transactions if t.category}
    if len(categories) > 1:
        add_sheet("Summary by Category", ("Category",), _summary.by_category(transactions))
    years = {t.assessment_year for t in transactions if t.assessment_year}
    if len(years) > 1:
        add_sheet("Summary by Year", ("Assessment Year",), _summary.by_year(transactions))
    add_sheet(
        "Summary by Deductor",
        ("Category", "Name of Deductor/Collector", "TAN"),
        _summary.by_deductor(transactions),
    )
    add_sheet("Summary by Section", ("Section",), _summary.by_section(transactions))
    add_sheet("Summary by Month", ("Month",), _summary.by_month(transactions))
