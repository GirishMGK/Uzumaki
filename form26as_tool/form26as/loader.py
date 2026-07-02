"""Load a Form 26AS file into a uniform grid of string cells.

Supported inputs: the native caret (^) delimited TRACES ``.txt`` export, a
genuine ``.xlsx``/``.xls`` workbook, an HTML export (including TRACES "Excel"
files that are really HTML with an ``.xls``/``.xlsx`` extension), and the
password-protected ``.pdf`` download. Whatever the source, this loader returns
a single flat grid (list of rows, each a list of stripped strings). The parser
scans that grid for the header rows it needs, so it does not matter that the
different parts of the statement have different column counts.
"""

from __future__ import annotations

from pathlib import Path
from typing import List, Optional

Grid = List[List[str]]


def _looks_like_html(head: bytes) -> bool:
    lowered = head.lower()
    return b"<html" in lowered or b"<table" in lowered or b"<!doctype html" in lowered


_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # legacy Excel (BIFF/.xls) container


def _looks_like_legacy_xls(head: bytes) -> bool:
    return head[:8] == _OLE2_MAGIC


def _load_xlsx(path: Path) -> Grid:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    grid: Grid = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            grid.append(["" if c is None else str(c).strip() for c in row])
    wb.close()
    return grid


def _load_text(path: Path) -> Grid:
    """Split a caret (^) delimited TRACES text export into a grid.

    Falls back to tab or comma if no carets are present.
    """
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        lines = fh.read().splitlines()
    sample = "\n".join(lines[:200])
    if "^" in sample:
        delim = "^"
    elif "\t" in sample:
        delim = "\t"
    else:
        delim = ","
    return [[cell.strip() for cell in line.split(delim)] for line in lines if line.strip()]


def _load_html(path: Path) -> Grid:
    from bs4 import BeautifulSoup

    with open(path, "rb") as fh:
        raw = fh.read()
    soup = BeautifulSoup(raw, "lxml")
    grid: Grid = []
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            cells: List[str] = []
            for cell in tr.find_all(["td", "th"]):
                text = cell.get_text(separator=" ", strip=True)
                # Expand colspan so columns stay aligned across rows: the first
                # slot holds the text, the rest are blank placeholders.
                try:
                    span = int(cell.get("colspan", 1))
                except (TypeError, ValueError):
                    span = 1
                cells.append(text)
                cells.extend([""] * max(0, span - 1))
            if cells:
                grid.append(cells)
    return grid


def _split_on_wide_gaps(line: str) -> List[str]:
    """Split a text line into cells on runs of 2+ spaces (fallback for PDFs)."""
    import re

    parts = re.split(r"\s{2,}", line.strip())
    return [p.strip() for p in parts if p.strip() != ""]


def _load_pdf(path: Path, password: Optional[str] = None) -> Grid:
    """Extract Part-A-style tables from a (possibly password-protected) PDF.

    Prefers pdfplumber's ruled-table extraction; if a page has no detectable
    table it falls back to splitting text lines on wide whitespace gaps.
    """
    import pdfplumber

    grid: Grid = []
    with pdfplumber.open(str(path), password=password or "") as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        grid.append(["" if c is None else str(c).replace("\n", " ").strip()
                                     for c in row])
            else:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    cells = _split_on_wide_gaps(line)
                    if cells:
                        grid.append(cells)
    return grid


def load_grid(path: str | Path, password: Optional[str] = None) -> Grid:
    """Return the file's contents as a list of rows of stripped string cells."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"No such file: {path}")

    with open(path, "rb") as fh:
        head = fh.read(4096)

    suffix = path.suffix.lower()
    if suffix in {".txt", ".csv"}:
        return _load_text(path)

    # PDF, by extension or magic bytes.
    if suffix == ".pdf" or head[:5] == b"%PDF-":
        return _load_pdf(path, password)

    # HTML masquerading as .xls/.xlsx is common from TRACES, so sniff content
    # first rather than trusting the extension.
    if _looks_like_html(head):
        return _load_html(path)

    # A caret-delimited text export sniffed by content (extension may vary).
    if b"^" in head and b"PK\x03\x04" not in head[:4]:
        return _load_text(path)

    if suffix in {".xlsx", ".xlsm", ".xls"}:
        try:
            return _load_xlsx(path)
        except Exception:
            if _looks_like_legacy_xls(head):
                # openpyxl can't read the pre-2007 binary format at all, and
                # feeding it to the HTML loader silently returns an empty
                # grid — surface a clear, actionable error instead.
                raise ValueError(
                    f"{path.name} is a legacy .xls (Excel 97-2003 binary) file, which isn't "
                    "supported. Open it in Excel and re-save as .xlsx, or export as CSV/text."
                ) from None
            # Genuine-looking extension but not a real workbook — fall back.
            return _load_html(path)
    if suffix in {".html", ".htm"}:
        return _load_html(path)

    # Unknown extension: try workbook, then HTML.
    try:
        return _load_xlsx(path)
    except Exception:
        if _looks_like_legacy_xls(head):
            raise ValueError(
                f"{path.name} is a legacy .xls (Excel 97-2003 binary) file, which isn't "
                "supported. Open it in Excel and re-save as .xlsx, or export as CSV/text."
            ) from None
        return _load_html(path)
