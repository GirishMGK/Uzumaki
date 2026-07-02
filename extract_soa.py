"""
SOA (Statement of Account) extractor for L&T Finance NBFC loan statements.

Parses the PDF SOA into a structured workbook with:
  - Loan_Master      : static loan & customer details (header block)
  - Finance_Summary  : "Loan Finance Summary" + receivable block
  - Disbursements    : disbursement summary table
  - Transactions     : full ledger (Date / Value Date / Particulars / DR / CR / Balance)
  - Charges          : discrete fee line-items (PF / BPI / insurance / TDS)
  - Bounce_Charge_Grid : the EMI-bounce charge slab table
  - Amortization     : reducing-balance schedule (recomputed)
  - TOC_TOD          : auto-computed audit validation checks

It also batches a folder of SOAs into a single PORTFOLIO EXCEPTION REPORT
(Portfolio_Summary / Exceptions / DPD_NPA / Charges) covering four check
families: EMI & amortization recompute, DPD/NPA staging, charges validation,
and ledger/balance reconciliation.

Usage:
    python extract_soa.py <input.pdf> [output.xlsx]          # single loan workbook
    python extract_soa.py --dir <folder> <output_folder>     # per-loan workbooks
    python extract_soa.py --portfolio <folder> [report.xlsx] # portfolio exception report
                                                             #   (+ per-loan books in ./loan_details)
"""
import re
import sys
import os
import datetime as dt

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MONTHS = "JAN FEB MAR APR MAY JUN JUL AUG SEP OCT NOV DEC".split()
DATE_RE = re.compile(r"^(\d{2}-[A-Z]{3}-\d{4})\s+(\d{2}-[A-Z]{3}-\d{4})\s+(.*)$")
NUM_RE = re.compile(r"-?[\d,]*\.\d{2}")


def to_num(s):
    if s is None:
        return None
    s = s.replace(",", "").strip()
    if s in ("", "NA", "N/A", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def to_date(s):
    try:
        return dt.datetime.strptime(s, "%d-%b-%Y").date()
    except (ValueError, TypeError):
        return None


def grab(text, label, stop=r"\n"):
    m = re.search(re.escape(label) + r"\s*:?\s*([^\n]+)", text)
    return m.group(1).strip() if m else None


_STOP_LABELS = (
    "Disbursement Date", "Loan Amount", "Annualised", "Interest Rate Type",
    "Interest Paid", "Loan Tenure", "Balance Tenure", "Loan Status",
    "Available Limit", "Moratorium", "First Instalment", "Instalment start",
    "Instalment End", "Frequency", "EMI Due Date", "Repayment Mode",
    "Principal Paid", "Instalment overdue", "Late Payment", "Excess amount",
    "Amount Under", "HSN Code", "Bank Account", "Bank Name", "Mandate Status",
    "Place of Supply", "State Code", "Customer GSTIN", "Customer ID",
    "Central KYC", "Email", "Mobile No", "Agreement No", "Currency",
)
STOP = r"(?=\s+(?:" + "|".join(_STOP_LABELS) + r")\b|\s*\n|\s*_|\s*$)"


def extract_master(text):
    fields = {
        "Applicant name": r"Applicant name:\s*(.+?)" + STOP,
        "Agreement No": r"Agreement No:\s*(.+?)" + STOP,
        "Customer ID": r"Customer ID:\s*(.+?)" + STOP,
        "Product": r"Product:\s*(.+?)" + STOP,
        "LTF GSTIN ID": r"LTF GSTIN ID:\s*(.+?)" + STOP,
        "Branch": r"Branch:\s*(.+?)" + STOP,
        "Disbursement Date": r"Disbursement Date:\s*([0-9A-Za-z\-]+)",
        "Loan Amount (Rs)": r"Loan Amount \(Rs\):\s*([\d,]+\.\d{2})",
        "Annualised Interest Rate %": r"Annualised Interest Rate %:\s*([\d.]+)",
        "Interest Rate Type": r"Interest Rate Type:\s*([A-Za-z]+)",
        "Loan Tenure (Months)": r"Loan Tenure \(Months\):\s*(\d+)",
        "Balance Tenure (Months)": r"Balance Tenure \(Months\):\s*(\d+)",
        "Loan Status": r"Loan Status:\s*([A-Za-z]+)",
        "Moratorium": r"Moratorium:\s*([A-Za-z]+)",
        "First Instalment Amount (Rs)": r"First Instalment Amount \(Rs\):\s*([\d,]+\.\d{2})",
        "Instalment start date": r"Instalment start date:\s*([0-9A-Za-z\-]+)",
        "Instalment End date": r"Instalment End date\s*:\s*([0-9A-Za-z\-]+)",
        "Frequency": r"Frequency:\s*([A-Za-z]+)",
        "EMI Due Date": r"EMI Due Date:\s*(.+?)" + STOP,
        "Repayment Mode": r"Repayment Mode:\s*([A-Za-z]+)",
        "Bank Name": r"Bank Name:\s*(.+?)" + STOP,
        "Bank Account": r"Bank Account:\s*(.+?)" + STOP,
        "Mandate Status": r"Mandate Status:\s*([A-Za-z]+)",
        "Principal Paid (Rs)": r"Principal Paid \(Rs\):\s*([\d,]+\.\d{2})",
        "Interest Paid (Rs)": r"Interest Paid \(Rs\):\s*([\d,]+\.\d{2})",
        "Instalment overdue (Rs)": r"Instalment overdue \(Rs\):\s*([\d,]+\.\d{2})",
        "Late Payment Charges (Rs)": r"Late Payment Charges \(Rs\):\s*([\d,]+\.\d{2})",
        "Excess amount": r"Excess amount:\s*([\d,]+\.\d{2})",
        "Amount Under Clearance": r"Amount Under Clearance:\s*([\d,]+\.\d{2})",
        "Mobile No.": r"Mobile No\.?:\s*(.+?)" + STOP,
        "Email": r"Email:\s*(\S+)",
        "State Code": r"State Code:\s*(\d+)",
        "Place of Supply": r"Place of Supply\s+(.+?)" + STOP,
        "Customer GSTIN ID": r"Customer GSTIN ID:\s*(\S+)",
        "HSN Code": r"HSN Code:\s*(\d+)",
        "Central KYC (CKYC) Id": r"Central KYC:\s*(\S+)",
        "Available Limit": r"Available Limit:\s*([\d,]*\.\d{2})",
        "Currency": r"Currency\s*:\s*([A-Z]+)",
    }
    out = {}
    for k, pat in fields.items():
        m = re.search(pat, text)
        if m:
            val = next((g for g in m.groups() if g), m.group(0))
            out[k] = val.strip()
        else:
            out[k] = None

    out["Customer Name"] = out.get("Applicant name")
    am = re.search(r"Customer Name & Contact Details Product Details\s*\n([^\n]+)", text)
    if am:
        out["Customer Name"] = am.group(1).strip()
    addr = re.findall(r"\n((?:\d+,\s*[^\n]+|[A-Z][A-Z ,]+ESTATE[^\n]*|NEAR[^\n]*))", text)
    if addr:
        out["Address / Contact Details"] = " ".join(a.strip() for a in addr[:4])

    m = re.search(r"for the period\s+([0-9A-Za-z\-]+)\s+to\s+([0-9A-Za-z\-]+)\s+Dated\s+([0-9A-Za-z\-]+)", text)
    if m:
        out["Statement From"], out["Statement To"], out["Statement Date"] = m.groups()
    return out


def extract_finance_summary(text):
    rows = []
    for label in ("Op.Bal", "Debits", "Credits", "Cl.Bal."):
        m = re.search(re.escape(label) + r"\.?\s+((?:-?[\d,]*\.\d{2}\s*){4,5})", text)
        if m:
            nums = NUM_RE.findall(m.group(1))
            rows.append([label] + nums)
    recv = {}
    m = re.search(r"Current OS Excess Receivable Accrued Interest Future Principal Total Receivable\s*\n\s*([^\n]+)", text)
    if m:
        nums = NUM_RE.findall(m.group(1))
        keys = ["Current OS", "Excess", "Receivable", "Accrued Interest", "Future Principal", "Total Receivable"]
        if len(nums) == 6:
            recv = dict(zip(keys, nums))
        elif len(nums) == 5:
            recv = dict(zip(["Current OS", "Excess Receivable", "Accrued Interest",
                             "Future Principal", "Total Receivable"], nums))
    return rows, recv


def extract_disbursements(text):
    rows = []
    m = re.search(r"DISBURSEMENT SUMMARY(.*?)(?:=====|$)", text, re.S)
    block = m.group(1) if m else text
    for line in block.splitlines():
        dm = re.match(r"\s*(\d+)\s+(\d{2}-[A-Z]{3}-\d{4})\s+([\d,]+\.\d{2})\s+(.*)", line)
        if dm:
            rows.append([int(dm.group(1)), dm.group(2), to_num(dm.group(3)), dm.group(4).strip()])
    return rows


def extract_bounce_summary(text):
    rows = []
    m = re.search(r"BOUNCE SUMMARY\s*\n.*?REASON\s*\n(.*?)(?:\n\s*The Company|BOUNCE CHARGES|=====|$)",
                  text, re.S)
    if not m:
        return rows
    for line in m.group(1).splitlines():
        line = line.strip()
        if not line or line.upper().startswith("NA NA"):
            continue
        dm = re.match(r"(\d{2}-[A-Z]{3}-\d{4})\s+(.*?)\s+([\d,]+\.\d{2})\s+(.*)$", line)
        if dm:
            rows.append([dm.group(1), dm.group(2).strip(), to_num(dm.group(3)), dm.group(4).strip()])
    return rows


def extract_part_payment(text):
    rows = []
    m = re.search(r"PART PAYMENT SUMMARY\s*\n[^\n]*Impact\s*\n(.*?)(?:\n\s*The amounts|BOUNCE SUMMARY|=====|$)",
                  text, re.S)
    if not m:
        return rows
    for line in m.group(1).splitlines():
        line = line.strip()
        if not line or line.upper().startswith("NA"):
            continue
        dm = re.match(r"(\d{2}-[A-Z]{3}-\d{4})\s+([\d,]+\.\d{2})\s+(.*)$", line)
        if dm:
            rows.append([dm.group(1), to_num(dm.group(2)), dm.group(3).strip()])
    return rows


def extract_transactions(pages_text):
    txns = []
    for text in pages_text:
        for line in text.splitlines():
            m = DATE_RE.match(line.strip())
            if not m:
                continue
            date, vdate, rest = m.groups()
            nums = NUM_RE.findall(rest)
            if len(nums) < 3:
                continue
            dr, cr, bal = nums[-3], nums[-2], nums[-1]
            particulars = rest
            for n in (bal, cr, dr):
                particulars = particulars.rsplit(n, 1)[0]
            particulars = particulars.strip()
            txns.append({
                "Date": date, "Value Date": vdate, "Particulars": particulars,
                "DR": to_num(dr), "CR": to_num(cr), "Balance": to_num(bal),
            })
    return txns


def emi_calc(principal, annual_rate, n):
    r = annual_rate / 12.0 / 100.0
    if r == 0:
        return principal / n
    return principal * r * (1 + r) ** n / ((1 + r) ** n - 1)


CRITICAL_FIELDS = [
    "Agreement No", "Loan Amount (Rs)", "Annualised Interest Rate %",
    "Loan Tenure (Months)", "First Instalment Amount (Rs)", "Disbursement Date",
    "Instalment start date", "Statement Date", "Principal Paid (Rs)",
    "Interest Paid (Rs)",
]


def assess_quality(master, fin_rows, recv, disb, txns):
    EMPTY = (None, "")
    missing_critical = [f for f in CRITICAL_FIELDS if master.get(f) in EMPTY]
    missing_other = [k for k, v in master.items()
                     if v in EMPTY and k not in CRITICAL_FIELDS]
    warnings = []
    notes = []
    if len(fin_rows) < 4:
        warnings.append("Loan Finance Summary not fully parsed")
    if recv.get("Total Receivable") in EMPTY:
        warnings.append("Receivable block not parsed")
    if not disb:
        warnings.append("No disbursement rows parsed")
    if not txns:
        warnings.append("No ledger transactions parsed")
    mangled = sum(1 for t in txns if re.match(r"^[\d,]+\.\d", t["Particulars"] or ""))
    if mangled:
        notes.append(f"{mangled} ledger row(s) with wrapped particulars (amounts OK)")

    captured = sum(1 for v in master.values() if v not in EMPTY)
    total = len(master) or 1
    status = "REVIEW" if (missing_critical or warnings) else "OK"
    return {
        "status": status,
        "missing_critical": missing_critical,
        "missing_other": missing_other,
        "warnings": warnings,
        "notes": notes,
        "captured": captured,
        "total": total,
        "completeness": round(captured / total * 100, 1),
    }


GST_RATE = 0.18
PF_PCT_THRESHOLD = 3.0
BPI_TOLERANCE = 0.15

CHARGE_MAP = [
    ("PROCESSING FEE", "Processing Fee"),
    ("BROKEN PERIOD INTEREST", "Broken Period Interest"),
    ("INSURANCE PREMIUM", "Insurance Premium"),
    ("TDS", "TDS"),
    ("BOUNCE", "Bounce / Penal Charge"),
    ("PENAL", "Bounce / Penal Charge"),
    ("LATE PAYMENT", "Late Payment Charge"),
    ("FORECLOSURE", "Foreclosure Charge"),
    ("PREPAYMENT", "Prepayment Charge"),
]


def extract_charges(txns):
    rows = []
    for t in txns:
        p = t["Particulars"].upper()
        if "DUE FOR INSTALMENT" in p or "AMT FINANCED" in p or "PMNT RCVD" in p:
            continue
        for kw, cat in CHARGE_MAP:
            if kw in p:
                rows.append({"Category": cat, "Date": t["Date"],
                             "Particulars": t["Particulars"],
                             "DR": t["DR"], "CR": t["CR"]})
                break
    return rows


def extract_bounce_grid(text):
    grid = []
    m = re.search(r"Charges in Rs Loan sanction amount in Rs\.\s*\n(.*?)(?:\n\s*[B-Z]\.|=====|$)",
                  text, re.S)
    if not m:
        return grid
    for line in m.group(1).splitlines():
        dm = re.match(r"\s*(\d+)\s+(.*)$", line.strip())
        if dm:
            grid.append({"charge": float(dm.group(1)), "slab": dm.group(2).strip()})
    return grid


def _lacs(text_amt):
    n = float(re.search(r"[\d.]+", text_amt).group())
    if "CR" in text_amt.upper():
        return n * 1e7
    if "LAC" in text_amt.upper():
        return n * 1e5
    return n


def expected_bounce_charge(grid, sanctioned):
    if not sanctioned:
        return None
    for g in grid:
        slab = g["slab"].upper()
        nums = re.findall(r"[\d.]+\s*(?:LACS?|CR\.?)", slab)
        if not nums:
            continue  # slab wording we don't recognise — skip rather than crash
        lo, hi = 0, float("inf")
        if slab.startswith("<"):
            hi = _lacs(nums[0])
        elif slab.startswith(">") and len(nums) == 1:
            lo = _lacs(nums[0])
        elif len(nums) >= 2:
            lo, hi = _lacs(nums[0]), _lacs(nums[1])
        elif len(nums) == 1:
            hi = _lacs(nums[0])
        if lo <= sanctioned <= hi or (hi == float("inf") and sanctioned >= lo):
            return g["charge"]
    return None


def build_amortization(principal, annual_rate, n, emi):
    sched = []
    r = annual_rate / 12.0 / 100.0
    bal = principal
    for i in range(1, int(n) + 1):
        interest = round(bal * r, 2)
        prin = round(emi - interest, 2)
        if i == int(n):
            prin = round(bal, 2)
            emi_i = round(prin + interest, 2)
        else:
            emi_i = emi
        bal = round(bal - prin, 2)
        sched.append({"No": i, "EMI": emi_i, "Interest": interest,
                      "Principal": prin, "Balance": max(bal, 0.0)})
    return sched


FREQ_MONTHS = {"MONTHLY": 1, "QUARTERLY": 3, "HALF": 6, "SEMI": 6,
               "ANNUAL": 12, "YEARLY": 12}


def _minus_months(d, months):
    m = d.month - 1 - months
    y = d.year + m // 12
    m = m % 12 + 1
    import calendar
    day = min(d.day, calendar.monthrange(y, m)[1])
    return dt.date(y, m, day)


def npa_stage(dpd):
    if dpd <= 0:
        return "Standard"
    if dpd <= 30:
        return "SMA-0"
    if dpd <= 60:
        return "SMA-1"
    if dpd <= 90:
        return "SMA-2"
    return "NPA (Sub-standard)"


def analyse_dpd(txns):
    dues = {}
    rows = []
    for t in txns:
        p = t["Particulars"].upper()
        m = re.search(r"DUE FOR INSTALMENT\s+(\d+)", p)
        if m and t["DR"]:
            dues[int(m.group(1))] = {"n": int(m.group(1)), "due_date": to_date(t["Date"]),
                                     "amount": t["DR"], "paid_date": None}
    pays = [t for t in txns if t["CR"] and ("CLEAR" in t["Particulars"].upper()
            or "PMNT RCVD" in t["Particulars"].upper() or "CHEQUE PMNT" in t["Particulars"].upper())]
    pay_idx = 0
    for n in sorted(dues):
        d = dues[n]
        for j in range(pay_idx, len(pays)):
            pd = to_date(pays[j]["Date"])
            if pd and d["due_date"] and pd >= d["due_date"]:
                d["paid_date"] = pd
                pay_idx = j + 1
                break
        dpd = (d["paid_date"] - d["due_date"]).days if d["paid_date"] and d["due_date"] else 0
        rows.append({"Instalment": n, "Due Date": d["due_date"], "Amount": d["amount"],
                     "Paid Date": d["paid_date"], "DPD": max(dpd, 0)})
    return rows


def build_checks(master, fin_rows, recv, disb, txns, charges=None,
                 bounce_grid=None, bounce_rows=None):
    checks = []
    charges = charges or []
    bounce_grid = bounce_grid or []
    bounce_rows = bounce_rows or []

    def add(cid, area, desc, expected, actual, status, remark=""):
        checks.append([cid, area, desc, expected, actual, status, remark])

    loan_amt = to_num(master.get("Loan Amount (Rs)"))
    rate = to_num(master.get("Annualised Interest Rate %"))
    tenure = to_num(master.get("Loan Tenure (Months)"))
    emi_actual = to_num(master.get("First Instalment Amount (Rs)"))
    prin_paid = to_num(master.get("Principal Paid (Rs)"))
    int_paid = to_num(master.get("Interest Paid (Rs)"))
    stmt_date = to_date(master.get("Statement Date") or "")
    disb_date = to_date(master.get("Disbursement Date") or "")
    inst_start = to_date(master.get("Instalment start date") or "")

    if loan_amt and rate and tenure and emi_actual:
        emi_exp = round(emi_calc(loan_amt, rate, int(tenure)), 2)
        diff = abs(emi_exp - emi_actual)
        st = "PASS" if diff <= max(50.0, emi_actual * 0.02) else "REVIEW"
        add("TOD-01", "Income/EMI", "Recompute EMI from sanctioned amount, rate & tenure",
            f"{emi_exp:,.2f}", f"{emi_actual:,.2f}", st,
            f"Diff {diff:,.2f}; EMI may differ for broken period / insurance funded")

    fut_prin = to_num(recv.get("Future Principal"))
    if loan_amt and prin_paid is not None and fut_prin is not None:
        tot = prin_paid + fut_prin
        st = "PASS" if abs(tot - loan_amt) <= 1.0 else "FAIL"
        add("TOD-02", "Principal", "Principal Paid + Future Principal = Loan Amount",
            f"{loan_amt:,.2f}", f"{tot:,.2f}", st, f"Diff {tot-loan_amt:,.2f}")

    fr = {r[0]: r for r in fin_rows}
    try:
        op = to_num(fr["Op.Bal"][-1]); dr = to_num(fr["Debits"][-1])
        cr = to_num(fr["Credits"][-1]); cl = to_num(fr["Cl.Bal."][-1])
        calc = op + dr - cr
        st = "PASS" if abs(calc - cl) <= 1.0 else "FAIL"
        add("TOD-03", "Ledger", "Op.Bal + Debits - Credits = Cl.Bal (Total)",
            f"{cl:,.2f}", f"{calc:,.2f}", st)
    except (KeyError, TypeError):
        pass

    comp = [to_num(recv.get(k)) for k in
            ("Current OS", "Excess", "Receivable", "Excess Receivable", "Accrued Interest", "Future Principal")]
    comp = [c for c in comp if c is not None]
    tot_recv = to_num(recv.get("Total Receivable"))
    if comp and tot_recv is not None:
        calc = sum(comp)
        st = "PASS" if abs(calc - tot_recv) <= 1.0 else "FAIL"
        add("TOD-04", "Receivable", "Sum of receivable components = Total Receivable",
            f"{tot_recv:,.2f}", f"{calc:,.2f}", st)

    if disb and loan_amt:
        d_tot = sum(d[2] for d in disb if d[2])
        st = "PASS" if abs(d_tot - loan_amt) <= 1.0 else "REVIEW"
        add("TOC-01", "Disbursement", "Sum of disbursals = Loan Amount (sanctioned)",
            f"{loan_amt:,.2f}", f"{d_tot:,.2f}", st)

    dpd_rows = analyse_dpd(txns)
    max_dpd = max((d["DPD"] for d in dpd_rows), default=0)
    st = ("PASS" if max_dpd <= 0 else "INFO" if max_dpd <= 30
          else "REVIEW" if max_dpd <= 90 else "FAIL")
    add("TOC-02", "DPD/NPA", "Max historical days-past-due across instalments",
        "0 days", f"{max_dpd} days ({npa_stage(max_dpd)})", st,
        "DPD>90 would have triggered NPA per RBI IRACP norms")

    cur_dpd = 0
    if stmt_date:
        unpaid = [d for d in dpd_rows if d["Paid Date"] is None
                  and d["Due Date"] and d["Due Date"] <= stmt_date]
        if unpaid:
            earliest = min(d["Due Date"] for d in unpaid)
            cur_dpd = (stmt_date - earliest).days
    stage = npa_stage(cur_dpd)
    st = "PASS" if cur_dpd <= 0 else ("REVIEW" if cur_dpd <= 90 else "FAIL")
    add("TOC-03", "DPD/NPA", "Current asset classification as on statement date",
        "Standard (0 DPD)", f"{cur_dpd} days -> {stage}", st)

    if int_paid is not None:
        st = "PASS" if int_paid >= 0 else "FAIL"
        add("TOC-04", "Income", "Interest Paid recorded (income recognition)",
            ">=0", f"{int_paid:,.2f}", st)

    pf_rows = [c for c in charges if c["Category"] == "Processing Fee" and c["DR"]]
    if pf_rows and loan_amt:
        incl = sum(c["DR"] for c in pf_rows)  # multi-disbursement loans can carry >1 PF line
        base = round(incl / (1 + GST_RATE), 2)
        pct = base / loan_amt * 100
        st = "PASS" if pct <= PF_PCT_THRESHOLD else "REVIEW"
        add("TOC-05", "Charges", "Processing fee % of sanctioned amount",
            f"<= {PF_PCT_THRESHOLD:.1f}%", f"{pct:.2f}% (base {base:,.2f} excl GST)", st,
            f"PF incl. tax {incl:,.2f}; assumed GST {GST_RATE*100:.0f}%")

    bpi_rows = [c for c in charges if c["Category"] == "Broken Period Interest" and c["DR"]]
    if bpi_rows and loan_amt and rate and disb_date and inst_start:
        months = FREQ_MONTHS.get((master.get("Frequency") or "MONTHLY").upper(), 1)
        cycle_start = _minus_months(inst_start, months)
        days = max((cycle_start - disb_date).days, 0)
        exp = round(loan_amt * rate / 100 * days / 365.0, 2)
        actual = sum(c["DR"] for c in bpi_rows)
        if exp:
            rel = abs(actual - exp) / exp
        else:
            # No broken-period interest is expected (disbursal fell on/after
            # the notional cycle start) — only PASS if none was actually
            # charged; a non-trivial actual amount is a real mismatch, not a
            # free pass.
            rel = 0.0 if abs(actual) <= 1 else float("inf")
        st = "PASS" if rel <= BPI_TOLERANCE else "REVIEW"
        add("TOD-06", "Charges", "Broken-period interest recompute (disbursal -> 1st billing cycle)",
            f"{exp:,.2f}", f"{actual:,.2f}", st,
            f"{days} days @ {rate:.2f}%; rel.diff {rel*100:.1f}%")

    exp_bounce = expected_bounce_charge(bounce_grid, loan_amt)
    if bounce_rows:
        for b in bounce_rows:
            actual = b[2]
            st = "PASS" if exp_bounce and abs((actual or 0) - exp_bounce) <= 1 else "REVIEW"
            add("TOC-06", "Charges", f"Bounce charge on {b[0]} vs grid",
                f"{exp_bounce:,.2f}" if exp_bounce else "per grid",
                f"{actual:,.2f}" if actual else "NA", st, b[1])
    elif exp_bounce is not None:
        add("TOC-06", "Charges", "Applicable bounce charge per grid (no bounces in period)",
            f"{exp_bounce:,.2f}", "No bounce", "INFO",
            "Grid rate for this sanction slab")

    if txns:
        last_bal = txns[-1]["Balance"]
        add("TOD-05", "Ledger", "Final running balance in transaction ledger",
            "informational", f"{last_bal:,.2f}", "INFO",
            "Should equal current overdue position")

    summary = {
        "Max DPD": max_dpd,
        "Current DPD": cur_dpd,
        "Current Stage": stage,
        "Exceptions": sum(1 for c in checks if c[5] in ("FAIL", "REVIEW")),
        "Fails": sum(1 for c in checks if c[5] == "FAIL"),
    }
    return checks, dpd_rows, summary


HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=12, color="1F4E78")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
STATUS_FILL = {"PASS": "C6EFCE", "FAIL": "FFC7CE", "REVIEW": "FFEB9C", "INFO": "DDEBF7"}


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, max_w=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(width + 3, max_w)


def write_workbook(path, master, fin_rows, recv, disb, txns, checks, dpd_rows,
                   part_pay=None, bounce=None, charges=None, amort=None, bounce_grid=None,
                   quality=None):
    wb = Workbook()

    ws = wb.active; ws.title = "Loan_Master"
    ws["A1"] = "LOAN MASTER & CUSTOMER DETAILS"; ws["A1"].font = TITLE_FONT
    ws.append([]); ws.append(["Field", "Value"]); style_header(ws, 3, 2)
    for k, v in master.items():
        ws.append([k, v])
    autosize(ws)

    ws = wb.create_sheet("Finance_Summary")
    ws["A1"] = "LOAN FINANCE SUMMARY"; ws["A1"].font = TITLE_FONT
    ws.append([])
    ws.append(["Particulars", "Principal", "Interest", "Late Payment Charges", "Misc. Charges", "Total"])
    style_header(ws, 3, 6)
    for r in fin_rows:
        ws.append([r[0]] + [to_num(x) for x in r[1:]])
    ws.append([])
    ws.append(["RECEIVABLE SUMMARY"]); ws.cell(ws.max_row, 1).font = TITLE_FONT
    ws.append(["Particulars", "Amount"]); style_header(ws, ws.max_row, 2)
    for k, v in recv.items():
        ws.append([k, to_num(v)])
    autosize(ws)

    ws = wb.create_sheet("Disbursements")
    ws.append(["Disbursal No", "Disbursal Date", "Disbursal Amount", "Particulars"])
    style_header(ws, 1, 4)
    for d in disb:
        ws.append(d)
    autosize(ws)

    ws = wb.create_sheet("Transactions")
    cols = ["Date", "Value Date", "Particulars", "DR", "CR", "Balance"]
    ws.append(cols); style_header(ws, 1, len(cols))
    for t in txns:
        ws.append([t["Date"], t["Value Date"], t["Particulars"], t["DR"], t["CR"], t["Balance"]])
    ws.freeze_panes = "A2"; autosize(ws)

    ws = wb.create_sheet("Part_Payment")
    ws.append(["Repayment Effective Date", "Part payment amount", "Impact"])
    style_header(ws, 1, 3)
    if part_pay:
        for p in part_pay:
            ws.append(p)
    else:
        ws.append(["NA", "NA", "NA"])
    autosize(ws)

    ws = wb.create_sheet("Bounce_Summary")
    ws.append(["Date of Bounce", "Narration", "Amount", "Reason"])
    style_header(ws, 1, 4)
    if bounce:
        for b in bounce:
            ws.append(b)
    else:
        ws.append(["NA", "NA", "NA", "NA"])
    autosize(ws)

    ws = wb.create_sheet("Charges")
    ws.append(["Category", "Date", "Particulars", "DR (Rs)", "CR (Rs)"])
    style_header(ws, 1, 5)
    for c in (charges or []):
        ws.append([c["Category"], c["Date"], c["Particulars"], c["DR"], c["CR"]])
    autosize(ws)

    ws = wb.create_sheet("Bounce_Charge_Grid")
    ws.append(["Charge (Rs)", "Loan sanction amount slab"])
    style_header(ws, 1, 2)
    for g in (bounce_grid or []):
        ws.append([g["charge"], g["slab"]])
    autosize(ws)

    ws = wb.create_sheet("Amortization")
    ws.append(["Instalment No", "EMI", "Interest", "Principal", "Closing Balance"])
    style_header(ws, 1, 5)
    for a in (amort or []):
        ws.append([a["No"], a["EMI"], a["Interest"], a["Principal"], a["Balance"]])
    ws.freeze_panes = "A2"; autosize(ws)

    ws = wb.create_sheet("DPD_Analysis")
    ws.append(["Instalment", "Due Date", "Amount", "Paid Date", "DPD (days)", "Stage"])
    style_header(ws, 1, 6)
    for d in dpd_rows:
        ws.append([d["Instalment"], str(d["Due Date"]), d["Amount"],
                   str(d["Paid Date"]), d["DPD"], npa_stage(d["DPD"])])
        if d["DPD"] > 0:
            ws.cell(ws.max_row, 5).fill = PatternFill("solid", fgColor="FFEB9C")
    autosize(ws)

    ws = wb.create_sheet("TOC_TOD")
    ws["A1"] = "TEST OF CONTROLS / TEST OF DETAILS - VALIDATION CHECKS"; ws["A1"].font = TITLE_FONT
    ws.append([])
    hdr = ["Check ID", "Area", "Procedure / Assertion", "Expected", "Actual", "Status", "Remark"]
    ws.append(hdr); style_header(ws, 3, len(hdr))
    for c in checks:
        ws.append(c)
        fill = STATUS_FILL.get(c[5])
        if fill:
            ws.cell(ws.max_row, 6).fill = PatternFill("solid", fgColor=fill)
    autosize(ws)

    if quality:
        ws = wb.create_sheet("Parse_Quality")
        ws["A1"] = "EXTRACTION QUALITY / DATA-INTEGRITY LOG"; ws["A1"].font = TITLE_FONT
        ws.append([])
        rows = [
            ("Parse Status", quality["status"]),
            ("Fields captured", f"{quality['captured']} / {quality['total']} ({quality['completeness']}%)"),
            ("Missing critical fields", ", ".join(quality["missing_critical"]) or "None"),
            ("Missing other fields", ", ".join(quality["missing_other"]) or "None"),
            ("Warnings (structural)", "; ".join(quality["warnings"]) or "None"),
            ("Notes (cosmetic)", "; ".join(quality.get("notes", [])) or "None"),
        ]
        ws.append(["Metric", "Value"]); style_header(ws, 3, 2)
        for k, v in rows:
            ws.append([k, v])
        sc = ws.cell(4, 2)
        sc.fill = PatternFill("solid", fgColor=STATUS_FILL.get(quality["status"], "DDEBF7"))
        autosize(ws)

    wb.save(path)


class NotAnSOAError(ValueError):
    pass


def extract_loan(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        pages = [p.extract_text() or "" for p in pdf.pages]
    full = "\n".join(pages)
    if re.search(r"Repayment Schedule of", full):
        raise NotAnSOAError("This is a Repayment Schedule (RPS), not a Statement of "
                            "Account. Use the RPS tab / extract_rps.py.")
    master = extract_master(full)
    fin_rows, recv = extract_finance_summary(full)
    disb = extract_disbursements(full)
    txns = extract_transactions(pages)
    part_pay = extract_part_payment(full)
    bounce = extract_bounce_summary(full)
    charges = extract_charges(txns)
    bounce_grid = extract_bounce_grid(full)

    loan_amt = to_num(master.get("Loan Amount (Rs)"))
    rate = to_num(master.get("Annualised Interest Rate %"))
    tenure = to_num(master.get("Loan Tenure (Months)"))
    emi = to_num(master.get("First Instalment Amount (Rs)"))
    # NOTE: use "is not None" rather than all(...) truthiness — a valid 0%
    # promotional loan has rate == 0.0, which all() would treat as missing.
    amort = (build_amortization(loan_amt, rate, tenure, emi)
             if None not in (loan_amt, rate, tenure, emi) else [])

    checks, dpd_rows, summary = build_checks(
        master, fin_rows, recv, disb, txns, charges, bounce_grid, bounce)
    quality = assess_quality(master, fin_rows, recv, disb, txns)

    return {
        "file": os.path.basename(pdf_path), "master": master, "fin_rows": fin_rows,
        "recv": recv, "disb": disb, "txns": txns, "part_pay": part_pay, "bounce": bounce,
        "charges": charges, "bounce_grid": bounce_grid, "amort": amort,
        "checks": checks, "dpd_rows": dpd_rows, "summary": summary, "quality": quality,
    }


def process(pdf_path, out_path):
    try:
        d = extract_loan(pdf_path)
    except NotAnSOAError:
        from extract_rps import extract_rps, write_rps_workbook
        print(f"[NOTICE] {os.path.basename(pdf_path)} is a Repayment Schedule - "
              f"producing the RPS workbook (Loan_Details + Repayment_Schedule).")
        ln = extract_rps(pdf_path)
        write_rps_workbook(out_path, [ln])
        print(f"[OK] {ln['file']} -> {out_path}  ({len(ln['schedule'])} schedule rows)")
        return None
    write_workbook(out_path, d["master"], d["fin_rows"], d["recv"], d["disb"], d["txns"],
                   d["checks"], d["dpd_rows"], part_pay=d["part_pay"], bounce=d["bounce"],
                   charges=d["charges"], amort=d["amort"], bounce_grid=d["bounce_grid"],
                   quality=d["quality"])
    print(f"[OK] {d['file']} -> {out_path}  ({len(d['txns'])} txns, "
          f"{len(d['checks'])} checks, {d['summary']['Exceptions']} exceptions)")
    return d


def write_portfolio(out_path, results, detail_dir=None):
    wb = Workbook()

    ws = wb.active; ws.title = "Portfolio_Summary"
    cols = ["File", "Agreement No", "Customer", "Product", "Sanctioned (Rs)", "Rate %",
            "Tenure", "Bal Tenure", "EMI (Rs)", "Disb Date", "Loan Status",
            "Principal Paid", "Interest Paid", "Overdue", "Current Stage",
            "Max DPD", "Exceptions", "Parse Status", "Result"]
    ws.append(cols); style_header(ws, 1, len(cols))
    for d in results:
        m, s, q = d["master"], d["summary"], d["quality"]
        result = "EXCEPTION" if s["Exceptions"] else "CLEAN"
        ws.append([d["file"], m.get("Agreement No"), m.get("Applicant name"),
                   m.get("Product"), to_num(m.get("Loan Amount (Rs)")),
                   to_num(m.get("Annualised Interest Rate %")),
                   to_num(m.get("Loan Tenure (Months)")), to_num(m.get("Balance Tenure (Months)")),
                   to_num(m.get("First Instalment Amount (Rs)")), m.get("Disbursement Date"),
                   m.get("Loan Status"), to_num(m.get("Principal Paid (Rs)")),
                   to_num(m.get("Interest Paid (Rs)")), to_num(m.get("Instalment overdue (Rs)")),
                   s["Current Stage"], s["Max DPD"], s["Exceptions"], q["status"], result])
        fill = "FFC7CE" if s["Fails"] else ("FFEB9C" if s["Exceptions"] else "C6EFCE")
        ws.cell(ws.max_row, len(cols)).fill = PatternFill("solid", fgColor=fill)
        ws.cell(ws.max_row, len(cols) - 1).fill = PatternFill(
            "solid", fgColor=STATUS_FILL.get(q["status"], "DDEBF7"))
    ws.freeze_panes = "A2"; autosize(ws)

    ws = wb.create_sheet("Exceptions")
    cols = ["File", "Agreement No", "Check ID", "Area", "Procedure / Assertion",
            "Expected", "Actual", "Status", "Remark"]
    ws.append(cols); style_header(ws, 1, len(cols))
    for d in results:
        agr = d["master"].get("Agreement No")
        for c in d["checks"]:
            if c[5] in ("FAIL", "REVIEW"):
                ws.append([d["file"], agr] + c)
                ws.cell(ws.max_row, 8).fill = PatternFill("solid", fgColor=STATUS_FILL.get(c[5]))
    ws.freeze_panes = "A2"; autosize(ws)

    ws = wb.create_sheet("DPD_NPA")
    cols = ["File", "Agreement No", "Instalment", "Due Date", "Amount",
            "Paid Date", "DPD (days)", "Stage"]
    ws.append(cols); style_header(ws, 1, len(cols))
    for d in results:
        agr = d["master"].get("Agreement No")
        for r in d["dpd_rows"]:
            if r["DPD"] > 0:
                ws.append([d["file"], agr, r["Instalment"], str(r["Due Date"]),
                           r["Amount"], str(r["Paid Date"]), r["DPD"], npa_stage(r["DPD"])])
    ws.freeze_panes = "A2"; autosize(ws)

    ws = wb.create_sheet("Charges")
    cols = ["File", "Agreement No", "Category", "Date", "Particulars", "DR (Rs)", "CR (Rs)"]
    ws.append(cols); style_header(ws, 1, len(cols))
    for d in results:
        agr = d["master"].get("Agreement No")
        for c in d["charges"]:
            ws.append([d["file"], agr, c["Category"], c["Date"], c["Particulars"], c["DR"], c["CR"]])
    ws.freeze_panes = "A2"; autosize(ws)

    ws = wb.create_sheet("Parse_Quality")
    cols = ["File", "Agreement No", "Parse Status", "Completeness %",
            "Missing Critical Fields", "Warnings", "Notes"]
    ws.append(cols); style_header(ws, 1, len(cols))
    for d in results:
        q = d["quality"]
        ws.append([d["file"], d["master"].get("Agreement No"), q["status"],
                   q["completeness"], ", ".join(q["missing_critical"]) or "None",
                   "; ".join(q["warnings"]) or "None",
                   "; ".join(q.get("notes", [])) or "None"])
        ws.cell(ws.max_row, 3).fill = PatternFill(
            "solid", fgColor=STATUS_FILL.get(q["status"], "DDEBF7"))
    ws.freeze_panes = "A2"; autosize(ws)

    wb.save(out_path)


def process_portfolio(folder, out_path, write_details=True):
    results = []
    detail_dir = os.path.join(os.path.dirname(out_path) or ".", "loan_details")
    if write_details:
        os.makedirs(detail_dir, exist_ok=True)
    for f in sorted(os.listdir(folder)):
        if not f.lower().endswith(".pdf"):
            continue
        path = os.path.join(folder, f)
        try:
            if write_details:
                d = process(path, os.path.join(detail_dir, os.path.splitext(f)[0] + ".xlsx"))
            else:
                d = extract_loan(path)
            if d:
                results.append(d)
        except Exception as e:
            print(f"[ERROR] {f}: {e}")
    if results:
        write_portfolio(out_path, results)
    return results


def main():
    args = sys.argv[1:]
    if not args:
        print(__doc__); sys.exit(1)
    if args[0] == "--portfolio":
        folder = args[1]
        out_path = args[2] if len(args) > 2 else "Portfolio_Exception_Report.xlsx"
        process_portfolio(folder, out_path)
    elif args[0] == "--dir":
        folder, out_folder = args[1], args[2]
        os.makedirs(out_folder, exist_ok=True)
        for f in os.listdir(folder):
            if f.lower().endswith(".pdf"):
                process(os.path.join(folder, f),
                        os.path.join(out_folder, os.path.splitext(f)[0] + ".xlsx"))
    else:
        pdf_path = args[0]
        out_path = args[1] if len(args) > 1 else os.path.splitext(pdf_path)[0] + ".xlsx"
        process(pdf_path, out_path)


if __name__ == "__main__":
    main()
