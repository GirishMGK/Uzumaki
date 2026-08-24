"""Builds seed/sample_masters.xlsx: a downloadable import template with a
Data Dictionary sheet, enum dropdowns, and 300 staff / 200 client sample
rows generated from the same pools as seed_data.py — so it doubles as a
"does the importer accept our own demo data cleanly" smoke fixture (see
tests/test_sample_workbook.py, P1 DoD in §13).

Run: cd backend && python -m seed.generate_sample_workbook
"""
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from app.models.enums import Designation, EmploymentStatus, EntityClass, RelationshipStatus, RiskRating, StaffCategory
from seed.seed_data import FIRST_NAMES, LAST_NAMES

random.seed(7)

OUT_PATH = Path(__file__).parent / "sample_masters.xlsx"

STAFF_COLUMNS = [
    ("employee_code", "Unique employee code", True, None),
    ("full_name", "Full name", True, None),
    ("short_name", "Preferred/short name", False, None),
    ("staff_category", "Category of staff", True, [e.value for e in StaffCategory]),
    ("designation", "Designation / title", True, [e.value for e in Designation]),
    ("grade_rank", "Seniority rank, 1 (most senior) - 12", True, None),
    ("employment_status", "Current employment status", False, [e.value for e in EmploymentStatus]),
    ("date_of_joining", "YYYY-MM-DD", False, None),
    ("official_email", "Firm email address", False, None),
    ("mobile", "Mobile number", False, None),
    ("home_city", "Home city", False, None),
]

CLIENT_COLUMNS = [
    ("client_code", "Unique client code", True, None),
    ("name", "Trading/short name", True, None),
    ("legal_name", "Full legal/registered name", False, None),
    ("entity_class", "Legal entity classification", True, [e.value for e in EntityClass]),
    ("relationship_status", "Firm relationship status", False, [e.value for e in RelationshipStatus]),
    ("risk_rating", "Audit risk rating", False, [e.value for e in RiskRating]),
    ("sector", "Industry sector", False, None),
    ("is_pie", "Public interest entity? TRUE/FALSE", False, ["TRUE", "FALSE"]),
]


def _add_dropdown(ws, col_index: int, allowed: list[str], max_row: int) -> None:
    formula = '"' + ",".join(allowed) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    col_letter = ws.cell(row=1, column=col_index).column_letter
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


def build() -> None:
    wb = Workbook()

    dd = wb.active
    dd.title = "Data Dictionary"
    dd.append(["Sheet", "Column", "Description", "Required", "Allowed values"])
    for sheet_name, columns in (("staff", STAFF_COLUMNS), ("clients", CLIENT_COLUMNS)):
        for col, desc, required, allowed in columns:
            dd.append([sheet_name, col, desc, "Yes" if required else "No", ", ".join(allowed) if allowed else ""])
    dd.freeze_panes = "A2"
    for i, width in enumerate([10, 22, 40, 10, 60], start=1):
        dd.column_dimensions[dd.cell(row=1, column=i).column_letter].width = width

    staff_ws = wb.create_sheet("staff")
    staff_ws.append([c[0] for c in STAFF_COLUMNS])
    designations_non_partner = [d for d in Designation if d != Designation.MANAGING_PARTNER]
    for i in range(300):
        fn, ln = random.choice(FIRST_NAMES), random.choice(LAST_NAMES)
        designation = random.choice(designations_non_partner)
        category = (
            StaffCategory.PARTNER if designation == Designation.PARTNER
            else StaffCategory.ARTICLED_ASSISTANT if "ARTICLE" in designation.value
            else StaffCategory.EMPLOYEE_CA
        )
        staff_ws.append([
            f"SAMP-E{i+1:04d}", f"{fn} {ln}", fn, category.value, designation.value,
            random.randint(2, 12), EmploymentStatus.ACTIVE.value, "2022-06-01",
            f"{fn.lower()}.{ln.lower()}{i}@firm.local", f"9{random.randint(100000000, 999999999)}",
            "Hyderabad",
        ])
    for idx, (col, _desc, _req, allowed) in enumerate(STAFF_COLUMNS, start=1):
        if allowed:
            _add_dropdown(staff_ws, idx, allowed, 301)
    staff_ws.freeze_panes = "A2"

    clients_ws = wb.create_sheet("clients")
    clients_ws.append([c[0] for c in CLIENT_COLUMNS])
    entity_classes = list(EntityClass)
    for i in range(200):
        clients_ws.append([
            f"SAMP-CL{i+1:04d}", f"{random.choice(LAST_NAMES)} {random.choice(['Industries', 'Enterprises', 'Ltd'])}",
            None, random.choice(entity_classes).value, RelationshipStatus.ACTIVE.value,
            random.choice(list(RiskRating)).value, random.choice(["BFSI", "Manufacturing", "IT", "Pharma"]),
            "TRUE" if i % 11 == 0 else "FALSE",
        ])
    for idx, (col, _desc, _req, allowed) in enumerate(CLIENT_COLUMNS, start=1):
        if allowed:
            _add_dropdown(clients_ws, idx, allowed, 201)
    clients_ws.freeze_panes = "A2"

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} (300 staff rows, 200 client rows)")


if __name__ == "__main__":
    build()
